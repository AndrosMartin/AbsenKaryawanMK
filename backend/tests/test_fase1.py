"""FASE 1 backend tests: change-password, employee creation (for manual-book test),
and report export with 'Detail Harian' sheet/page."""
import io
import os
import uuid
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://office-monitor-2.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

OWNER_EMAIL = "owner@company.com"
OWNER_PASS = "password123"


def _login(email, password):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=30)
    assert r.status_code == 200, f"login failed {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def owner_token():
    return _login(OWNER_EMAIL, OWNER_PASS)


@pytest.fixture(scope="session")
def staff_user(owner_token):
    """Create a throwaway staff user via owner."""
    email = f"test_staff_{uuid.uuid4().hex[:8]}@company.com"
    pwd = "password123"
    headers = {"Authorization": f"Bearer {owner_token}"}
    body = {"name": "TEST Staff Fase1", "email": email, "password": pwd,
            "role": "staff", "department": "QA", "position": "Tester"}
    r = requests.post(f"{API}/employees", json=body, headers=headers, timeout=30)
    assert r.status_code == 200, f"employee create failed: {r.status_code} {r.text}"
    data = r.json()
    yield {"email": email, "password": pwd, "id": data.get("id")}
    # cleanup
    if data.get("id"):
        requests.delete(f"{API}/employees/{data['id']}", headers=headers, timeout=30)


class TestChangePassword:
    def test_wrong_current_password(self, staff_user):
        token = _login(staff_user["email"], staff_user["password"])
        r = requests.post(f"{API}/auth/change-password",
                          json={"current_password": "WRONG", "new_password": "newpass123"},
                          headers={"Authorization": f"Bearer {token}"}, timeout=30)
        assert r.status_code == 400
        assert "Password lama salah" in r.json().get("detail", "")

    def test_new_password_too_short(self, staff_user):
        token = _login(staff_user["email"], staff_user["password"])
        r = requests.post(f"{API}/auth/change-password",
                          json={"current_password": staff_user["password"], "new_password": "abc"},
                          headers={"Authorization": f"Bearer {token}"}, timeout=30)
        assert r.status_code == 400
        assert "minimal 6" in r.json().get("detail", "")

    def test_new_same_as_current(self, staff_user):
        token = _login(staff_user["email"], staff_user["password"])
        r = requests.post(f"{API}/auth/change-password",
                          json={"current_password": staff_user["password"],
                                "new_password": staff_user["password"]},
                          headers={"Authorization": f"Bearer {token}"}, timeout=30)
        assert r.status_code == 400
        assert "berbeda" in r.json().get("detail", "").lower()

    def test_valid_change_then_login(self, staff_user):
        token = _login(staff_user["email"], staff_user["password"])
        new_pwd = "newpass456"
        r = requests.post(f"{API}/auth/change-password",
                          json={"current_password": staff_user["password"], "new_password": new_pwd},
                          headers={"Authorization": f"Bearer {token}"}, timeout=30)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("ok") is True
        # login with new password
        token2 = _login(staff_user["email"], new_pwd)
        assert token2
        # revert so subsequent test fixtures still work
        requests.post(f"{API}/auth/change-password",
                      json={"current_password": new_pwd, "new_password": staff_user["password"]},
                      headers={"Authorization": f"Bearer {token2}"}, timeout=30)


class TestExportSummaryDetail:
    def _range(self):
        # past 30 days
        from datetime import datetime, timedelta, timezone
        today = datetime.now(timezone.utc).date()
        start = (today - timedelta(days=30)).isoformat()
        end = today.isoformat()
        return start, end

    def test_xlsx_has_two_sheets_with_correct_columns(self, owner_token):
        start, end = self._range()
        r = requests.get(f"{API}/attendance/summary/export",
                         params={"format": "xlsx", "start": start, "end": end},
                         headers={"Authorization": f"Bearer {owner_token}"}, timeout=60)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames == ["Rekap Absensi", "Detail Harian"], wb.sheetnames
        ws2 = wb["Detail Harian"]
        # row 4 is the header row
        headers = [c.value for c in ws2[4]]
        expected = ["Tanggal", "ID Karyawan", "Nama", "Departemen",
                    "Jam Masuk", "Jam Pulang", "Status"]
        assert headers == expected, headers

    def test_pdf_export_valid(self, owner_token):
        start, end = self._range()
        r = requests.get(f"{API}/attendance/summary/export",
                         params={"format": "pdf", "start": start, "end": end},
                         headers={"Authorization": f"Bearer {owner_token}"}, timeout=60)
        assert r.status_code == 200, r.text
        assert r.headers.get("content-type") == "application/pdf"
        assert r.content[:4] == b"%PDF"


class TestRegression:
    def test_owner_login_me(self, owner_token):
        r = requests.get(f"{API}/auth/me",
                         headers={"Authorization": f"Bearer {owner_token}"}, timeout=30)
        assert r.status_code == 200
        assert r.json().get("role") == "owner"

    def test_dashboard_stats(self, owner_token):
        r = requests.get(f"{API}/dashboard/stats",
                         headers={"Authorization": f"Bearer {owner_token}"}, timeout=30)
        assert r.status_code == 200
        d = r.json()
        for k in ("total", "present", "late", "absent", "departments", "trend"):
            assert k in d
