from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Query
from fastapi.responses import Response
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from pydantic import BaseModel, Field, EmailStr, ConfigDict, BeforeValidator
from typing import List, Optional, Annotated, Any
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
import logging
import math
import uuid
import bcrypt
import jwt
import secrets
import json
import asyncio
import io
from pywebpush import webpush, WebPushException
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

# ----------------------------------------------------------------------------
# Setup
# ----------------------------------------------------------------------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
TZ = ZoneInfo("Asia/Jakarta")
WORK_START = os.environ.get("WORK_START", "09:00")
FACE_MATCH_THRESHOLD = 0.55
SEED_DEMO = os.environ.get("SEED_DEMO", "true").strip().lower() == "true"
VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY", "")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_SUBJECT = os.environ.get("VAPID_SUBJECT", "mailto:admin@example.com")
REMINDER_TIME = os.environ.get("REMINDER_TIME", "08:00")
scheduler = AsyncIOScheduler(timezone=TZ)

MONITOR_ROLES = {"owner", "direksi", "manager", "hrd"}
MANAGE_ROLES = {"owner", "direksi"}        # dapat menerapkan perubahan langsung
HR_ROLES = {"owner", "direksi", "hrd"}     # dapat mengajukan CRUD karyawan
APPROVER_ROLES = {"owner", "direksi"}      # dapat menyetujui/menolak permintaan HRD
VALID_ROLES = {"owner", "direksi", "manager", "hrd", "staff"}
HRD_ASSIGNABLE_ROLES = {"manager", "hrd", "staff"}  # role yang boleh diatur HRD

app = FastAPI(title="AbsensiPro API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("absensipro")

PyObjectId = Annotated[str, BeforeValidator(str)]


# ----------------------------------------------------------------------------
# Auth helpers
# ----------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


def require_roles(roles: set):
    async def checker(user: dict = Depends(get_current_user)) -> dict:
        if user.get("role") not in roles:
            raise HTTPException(status_code=403, detail="Akses ditolak untuk role Anda")
        return user
    return checker


def public_user(u: dict) -> dict:
    return {
        "id": str(u["_id"]),
        "email": u["email"],
        "name": u.get("name"),
        "role": u.get("role"),
        "department": u.get("department"),
        "position": u.get("position"),
        "employee_id": u.get("employee_id"),
        "phone": u.get("phone"),
        "is_reviewer": bool(u.get("is_reviewer")),
        "face_enrolled": bool(u.get("face_descriptor")),
        "created_at": u.get("created_at"),
    }


# ----------------------------------------------------------------------------
# Utilities
# ----------------------------------------------------------------------------
def haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def euclidean(a: List[float], b: List[float]) -> float:
    return math.sqrt(sum((x - y) ** 2 for x, y in zip(a, b)))


def today_str() -> str:
    return datetime.now(TZ).strftime("%Y-%m-%d")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _oid(s: str) -> ObjectId:
    try:
        return ObjectId(s)
    except Exception:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")


DEFAULT_SETTINGS = {"work_start": "09:00", "work_end": "17:00", "tolerance_minutes": 15,
                    "leave_quota": 12}


async def get_settings() -> dict:
    doc = await db.app_settings.find_one({"_id": "work_schedule"})
    if not doc:
        return dict(DEFAULT_SETTINGS)
    return {
        "work_start": doc.get("work_start", DEFAULT_SETTINGS["work_start"]),
        "work_end": doc.get("work_end", DEFAULT_SETTINGS["work_end"]),
        "tolerance_minutes": int(doc.get("tolerance_minutes", DEFAULT_SETTINGS["tolerance_minutes"])),
        "leave_quota": int(doc.get("leave_quota", DEFAULT_SETTINGS["leave_quota"])),
    }


async def _apply_settings(payload: dict) -> None:
    await db.app_settings.update_one({"_id": "work_schedule"}, {"$set": payload}, upsert=True)


def _validate_hhmm(s: str) -> None:
    try:
        hh, mm = map(int, s.split(":"))
        assert 0 <= hh < 24 and 0 <= mm < 60
    except Exception:
        raise HTTPException(status_code=400, detail="Format jam tidak valid (HH:MM)")


def compute_status(check_in_dt: datetime, work_start: str = "09:00",
                   tolerance_minutes: int = 15) -> str:
    local = check_in_dt.astimezone(TZ)
    hh, mm = map(int, work_start.split(":"))
    limit = local.replace(hour=hh, minute=mm, second=0, microsecond=0)
    grace = limit + timedelta(minutes=tolerance_minutes)
    if local <= limit:
        return "present"
    if local <= grace:
        return "tolerance"
    return "late"


# ----------------------------------------------------------------------------
# Web Push (VAPID)
# ----------------------------------------------------------------------------
def _send_one_push(sub: dict, payload: str) -> None:
    webpush(subscription_info=sub, data=payload,
            vapid_private_key=VAPID_PRIVATE_KEY,
            vapid_claims={"sub": VAPID_SUBJECT})


async def send_push_to_users(user_ids, title: str, body: str,
                             url: str = "/", tag: str = "absensipro") -> None:
    if not VAPID_PRIVATE_KEY or not user_ids:
        return
    uids = list({str(u) for u in user_ids})
    subs = await db.push_subscriptions.find({"user_id": {"$in": uids}}).to_list(2000)
    if not subs:
        return
    payload = json.dumps({"title": title, "body": body, "url": url, "tag": tag})
    loop = asyncio.get_running_loop()
    for s in subs:
        try:
            await loop.run_in_executor(None, _send_one_push, s["subscription"], payload)
        except WebPushException as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            if status in (404, 410):
                await db.push_subscriptions.delete_one({"_id": s["_id"]})
            else:
                logger.warning(f"web push failed: {e}")
        except Exception as e:
            logger.warning(f"web push error: {e}")


async def _approver_ids() -> list:
    docs = await db.users.find({"role": {"$in": list(APPROVER_ROLES)}}, {"_id": 1}).to_list(200)
    return [str(d["_id"]) for d in docs]


async def _send_attendance_reminders() -> None:
    """Pengingat harian: kirim ke karyawan yang belum check-in hari ini."""
    date = today_str()
    users = await db.users.find({}, {"_id": 1, "name": 1}).to_list(2000)
    for u in users:
        rec = await db.attendance.find_one(
            {"user_id": str(u["_id"]), "date": date, "check_in": {"$ne": None}})
        if rec:
            continue
        await send_push_to_users(
            [str(u["_id"])], "Pengingat Absensi",
            "Jangan lupa check-in hari ini. Selamat bekerja!",
            url="/#checkin", tag="reminder")


# ----------------------------------------------------------------------------
# Models
# ----------------------------------------------------------------------------
class RegisterBody(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str = "staff"
    department: Optional[str] = "Umum"
    position: Optional[str] = "Staff"
    phone: Optional[str] = None


class LoginBody(BaseModel):
    email: EmailStr
    password: str


class ChangePasswordBody(BaseModel):
    current_password: str
    new_password: str


class SettingsBody(BaseModel):
    work_start: str
    work_end: str
    tolerance_minutes: int
    leave_quota: Optional[int] = 12


class LeaveRequestBody(BaseModel):
    leave_type: str = "tahunan"
    start_date: str
    end_date: str
    reason: Optional[str] = ""


class LeaveDecisionBody(BaseModel):
    note: Optional[str] = ""


class ReviewerBody(BaseModel):
    is_reviewer: bool


class EmployeeBody(BaseModel):
    name: str
    email: EmailStr
    password: str = "password123"
    role: str = "staff"
    department: Optional[str] = "Umum"
    position: Optional[str] = "Staff"
    phone: Optional[str] = None


class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    password: Optional[str] = None


class FaceEnrollBody(BaseModel):
    descriptor: List[float]


class OfficeBody(BaseModel):
    name: str
    lat: float
    lng: float
    radius_m: int = 200


class CheckInBody(BaseModel):
    method: str  # "face" | "qr"
    lat: float
    lng: float
    descriptor: Optional[List[float]] = None
    qr_code: Optional[str] = None


class CheckOutBody(BaseModel):
    lat: Optional[float] = None
    lng: Optional[float] = None


class RejectBody(BaseModel):
    reason: Optional[str] = None


class PushSubscribeBody(BaseModel):
    subscription: dict


class PushUnsubscribeBody(BaseModel):
    endpoint: Optional[str] = None


# ----------------------------------------------------------------------------
# Auth routes
# ----------------------------------------------------------------------------
@api_router.post("/auth/register")
async def register(body: RegisterBody):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    count = await db.users.count_documents({})
    doc = {
        "name": body.name, "email": email,
        "password_hash": hash_password(body.password),
        "role": body.role if body.role in VALID_ROLES else "staff",
        "department": body.department, "position": body.position,
        "phone": body.phone,
        "employee_id": f"EMP-{count + 1:04d}",
        "face_descriptor": None,
        "created_at": now_iso(),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    token = create_access_token(str(res.inserted_id), email, doc["role"])
    return {"token": token, "user": public_user(doc)}


@api_router.post("/auth/login")
async def login(body: LoginBody):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    token = create_access_token(str(user["_id"]), email, user["role"])
    return {"token": token, "user": public_user(user)}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return public_user(user)


@api_router.post("/auth/change-password")
async def change_password(body: ChangePasswordBody, user: dict = Depends(get_current_user)):
    if not verify_password(body.current_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Password lama salah")
    if len(body.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password baru minimal 6 karakter")
    if body.new_password == body.current_password:
        raise HTTPException(status_code=400, detail="Password baru harus berbeda dari password lama")
    await db.users.update_one({"_id": user["_id"]},
                              {"$set": {"password_hash": hash_password(body.new_password)}})
    return {"ok": True, "message": "Password berhasil diubah"}


# ----------------------------------------------------------------------------
# Face enrollment
# ----------------------------------------------------------------------------
@api_router.post("/face/enroll")
async def enroll_face(body: FaceEnrollBody, user: dict = Depends(get_current_user)):
    if len(body.descriptor) != 128:
        raise HTTPException(status_code=400, detail="Descriptor wajah tidak valid")
    await db.users.update_one({"_id": user["_id"]},
                              {"$set": {"face_descriptor": body.descriptor}})
    return {"ok": True, "message": "Wajah berhasil didaftarkan"}


# ----------------------------------------------------------------------------
# Offices
# ----------------------------------------------------------------------------
@api_router.get("/offices")
async def list_offices(user: dict = Depends(get_current_user)):
    offices = await db.offices.find().to_list(100)
    return [{"id": str(o["_id"]), "name": o["name"], "lat": o["lat"],
             "lng": o["lng"], "radius_m": o["radius_m"], "qr_code": o["qr_code"]}
            for o in offices]


@api_router.post("/offices")
async def create_office(body: OfficeBody, user: dict = Depends(require_roles(MANAGE_ROLES))):
    doc = body.model_dump()
    doc["qr_code"] = f"OFFICE-{uuid.uuid4().hex[:10].upper()}"
    res = await db.offices.insert_one(doc)
    doc["_id"] = res.inserted_id
    return {"id": str(res.inserted_id), **{k: doc[k] for k in ("name", "lat", "lng", "radius_m", "qr_code")}}


@api_router.put("/offices/{office_id}")
async def update_office(office_id: str, body: OfficeBody, user: dict = Depends(require_roles(MANAGE_ROLES))):
    await db.offices.update_one({"_id": ObjectId(office_id)}, {"$set": body.model_dump()})
    return {"ok": True}


@api_router.delete("/offices/{office_id}")
async def delete_office(office_id: str, user: dict = Depends(require_roles(MANAGE_ROLES))):
    await db.offices.delete_one({"_id": ObjectId(office_id)})
    return {"ok": True}


async def validate_geofence(lat: float, lng: float):
    offices = await db.offices.find().to_list(100)
    if not offices:
        raise HTTPException(status_code=400, detail="Belum ada lokasi kantor yang dikonfigurasi")
    best = None
    for o in offices:
        d = haversine_m(lat, lng, o["lat"], o["lng"])
        if best is None or d < best[1]:
            best = (o, d)
    office, dist = best
    within = dist <= office["radius_m"]
    return office, dist, within


def verify_attendance_identity(user: dict, body: "CheckInBody", office: dict) -> None:
    """Raise HTTPException if the chosen verification method fails."""
    if body.method == "face":
        stored = user.get("face_descriptor")
        if not stored:
            raise HTTPException(status_code=400, detail="Wajah belum didaftarkan. Daftarkan wajah Anda dulu.")
        if not body.descriptor or len(body.descriptor) != 128:
            raise HTTPException(status_code=400, detail="Data wajah tidak terbaca, coba lagi")
        if euclidean(stored, body.descriptor) > FACE_MATCH_THRESHOLD:
            raise HTTPException(status_code=400, detail="Wajah tidak cocok dengan data terdaftar")
    elif body.method == "qr":
        if not body.qr_code or body.qr_code != office["qr_code"]:
            raise HTTPException(status_code=400, detail="QR Code tidak valid untuk lokasi ini")
    else:
        raise HTTPException(status_code=400, detail="Metode absensi tidak dikenal")


# ----------------------------------------------------------------------------
# Attendance
# ----------------------------------------------------------------------------
def att_public(a: dict, user: Optional[dict] = None) -> dict:
    out = {
        "id": str(a["_id"]),
        "user_id": a["user_id"],
        "date": a["date"],
        "check_in": a.get("check_in"),
        "check_out": a.get("check_out"),
        "status": a.get("status"),
        "method": a.get("method"),
        "office_name": a.get("office_name"),
        "distance_m": a.get("distance_m"),
        "within_geofence": a.get("within_geofence"),
    }
    if user:
        out.update({"name": user.get("name"), "department": user.get("department"),
                    "position": user.get("position"), "employee_id": user.get("employee_id"),
                    "role": user.get("role")})
    return out


@api_router.post("/attendance/check-in")
async def check_in(body: CheckInBody, user: dict = Depends(get_current_user)):
    date = today_str()
    existing = await db.attendance.find_one({"user_id": str(user["_id"]), "date": date})
    if existing and existing.get("check_in"):
        raise HTTPException(status_code=400, detail="Anda sudah melakukan check-in hari ini")

    office, dist, within = await validate_geofence(body.lat, body.lng)
    if not within:
        raise HTTPException(status_code=400,
                            detail=f"Anda berada {int(dist)}m dari {office['name']} (radius {office['radius_m']}m). Mendekatlah ke kantor.")

    verify_attendance_identity(user, body, office)

    check_in_dt = datetime.now(timezone.utc)
    settings = await get_settings()
    status = compute_status(check_in_dt, settings["work_start"], settings["tolerance_minutes"])
    doc = {
        "user_id": str(user["_id"]),
        "date": date,
        "check_in": check_in_dt.isoformat(),
        "check_out": None,
        "status": status,
        "method": body.method,
        "office_name": office["name"],
        "distance_m": int(dist),
        "within_geofence": True,
        "lat": body.lat, "lng": body.lng,
    }
    await db.attendance.insert_one(doc)
    if status == "late":
        t = check_in_dt.astimezone(TZ).strftime("%H:%M")
        await send_push_to_users([str(user["_id"])], "Absensi Terlambat",
                                 f"Anda tercatat TELAT pada {t} WIB.",
                                 url="/#history", tag="late")
        await send_push_to_users(await _approver_ids(), "Karyawan Terlambat",
                                 f"{user.get('name')} check-in telat ({t} WIB).",
                                 url="/#monitoring", tag="late")
    doc["_id"] = "tmp"
    return att_public(doc, user)


@api_router.post("/attendance/check-out")
async def check_out(body: CheckOutBody, user: dict = Depends(get_current_user)):
    date = today_str()
    rec = await db.attendance.find_one({"user_id": str(user["_id"]), "date": date})
    if not rec or not rec.get("check_in"):
        raise HTTPException(status_code=400, detail="Anda belum check-in hari ini")
    if rec.get("check_out"):
        raise HTTPException(status_code=400, detail="Anda sudah check-out hari ini")
    await db.attendance.update_one({"_id": rec["_id"]},
                                   {"$set": {"check_out": now_iso()}})
    rec["check_out"] = now_iso()
    return att_public(rec, user)


@api_router.get("/attendance/today")
async def attendance_today(user: dict = Depends(get_current_user)):
    rec = await db.attendance.find_one({"user_id": str(user["_id"]), "date": today_str()})
    if not rec:
        return {"checked_in": False}
    return {"checked_in": True, **att_public(rec, user)}


@api_router.get("/attendance/me")
async def my_attendance(user: dict = Depends(get_current_user), limit: int = 60):
    recs = await db.attendance.find({"user_id": str(user["_id"])}).sort("date", -1).to_list(limit)
    return [att_public(r, user) for r in recs]


@api_router.get("/attendance")
async def all_attendance(date: Optional[str] = Query(None),
                         user: dict = Depends(require_roles(MONITOR_ROLES))):
    date = date or today_str()
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    recs = await db.attendance.find({"date": date}).to_list(2000)
    rec_by_user = {r["user_id"]: r for r in recs}
    rows = []
    for u in users:
        uid = str(u["_id"])
        r = rec_by_user.get(uid)
        if r:
            rows.append(att_public(r, u))
        else:
            rows.append({"id": None, "user_id": uid, "date": date, "check_in": None,
                         "check_out": None, "status": "absent", "method": None,
                         "office_name": None, "distance_m": None, "within_geofence": None,
                         "name": u.get("name"), "department": u.get("department"),
                         "position": u.get("position"), "employee_id": u.get("employee_id"),
                         "role": u.get("role")})
    return rows


# ----------------------------------------------------------------------------
# Attendance summary (rekap) + export PDF/Excel
# ----------------------------------------------------------------------------
_STATUS_LABEL = {"present": "Tepat Waktu", "tolerance": "Toleransi", "late": "Terlambat",
                 "absent": "Tidak Hadir"}


def _fmt_local_time(iso) -> str:
    if not iso:
        return "-"
    try:
        dt = datetime.fromisoformat(iso)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(TZ).strftime("%H:%M")
    except Exception:
        return "-"


def _count_workdays(start: str, end: str) -> int:
    """Jumlah hari kerja (Senin–Jumat) dalam rentang, dibatasi s/d hari ini."""
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end, "%Y-%m-%d").date()
    today = datetime.now(TZ).date()
    if e > today:
        e = today
    n, d = 0, s
    while d <= e:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


async def _build_summary(start: str, end: str,
                         user_id: Optional[str] = None, q: Optional[str] = None,
                         include_detail: bool = False) -> dict:
    users = await db.users.find({}, {"password_hash": 0}).sort("name", 1).to_list(1000)
    if user_id:
        users = [u for u in users if str(u["_id"]) == user_id]
    if q:
        ql = q.lower()
        users = [u for u in users
                 if ql in ((u.get("name") or "") + " " + (u.get("department") or "")).lower()]
    uids = [str(u["_id"]) for u in users]
    recs = await db.attendance.find(
        {"date": {"$gte": start, "$lte": end}, "user_id": {"$in": uids}}).to_list(50000)
    workdays = _count_workdays(start, end)
    agg = {}
    for r in recs:
        a = agg.setdefault(r["user_id"], {"present": 0, "tolerance": 0, "late": 0, "attended": 0})
        st = r.get("status")
        if st == "present":
            a["present"] += 1
        elif st == "tolerance":
            a["tolerance"] += 1
        elif st == "late":
            a["late"] += 1
        if r.get("check_in"):
            a["attended"] += 1
    rows, tot = [], {"present": 0, "tolerance": 0, "late": 0, "absent": 0, "attended": 0}
    for u in users:
        a = agg.get(str(u["_id"]), {"present": 0, "tolerance": 0, "late": 0, "attended": 0})
        attended = a["attended"]
        absent = max(0, workdays - attended)
        rate = round(attended / workdays * 100, 1) if workdays else 0
        rows.append({
            "user_id": str(u["_id"]), "employee_id": u.get("employee_id"),
            "name": u.get("name"), "department": u.get("department"),
            "position": u.get("position"), "role": u.get("role"),
            "present": a["present"], "tolerance": a["tolerance"], "late": a["late"],
            "absent": absent, "attended": attended, "workdays": workdays, "rate": rate,
        })
        tot["present"] += a["present"]
        tot["tolerance"] += a["tolerance"]
        tot["late"] += a["late"]
        tot["attended"] += attended
        tot["absent"] += absent
    out = {"start": start, "end": end, "workdays": workdays, "rows": rows, "totals": tot}
    if include_detail:
        user_by_id = {str(u["_id"]): u for u in users}
        detail = []
        for r in recs:
            if not r.get("check_in"):
                continue
            u = user_by_id.get(r["user_id"]) or {}
            detail.append({
                "date": r.get("date"),
                "employee_id": u.get("employee_id"),
                "name": u.get("name"),
                "department": u.get("department"),
                "check_in": _fmt_local_time(r.get("check_in")),
                "check_out": _fmt_local_time(r.get("check_out")),
                "status": r.get("status"),
                "status_label": _STATUS_LABEL.get(r.get("status"), r.get("status") or "-"),
            })
        detail.sort(key=lambda x: (x["date"] or "", x["name"] or ""))
        out["detail"] = detail
    return out


@api_router.get("/attendance/summary")
async def attendance_summary(start: str = Query(...), end: str = Query(...),
                             user_id: Optional[str] = Query(None), q: Optional[str] = Query(None),
                             user: dict = Depends(require_roles(MONITOR_ROLES))):
    return await _build_summary(start, end, user_id, q)


def _latin(s) -> str:
    return str(s or "").encode("latin-1", "replace").decode("latin-1")


def _xlsx_bytes(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Absensi"
    ws.append(["Rekap Absensi Karyawan — MitraKeuangan"])
    ws.append([f"Periode: {data['start']} s/d {data['end']}    Hari kerja: {data['workdays']}"])
    ws.append([])
    headers = ["ID Karyawan", "Nama", "Departemen", "Posisi", "Tepat Waktu", "Toleransi",
               "Terlambat", "Tidak Hadir", "Total Hadir", "Hari Kerja", "Kehadiran (%)"]
    ws.append(headers)
    head_row = ws.max_row
    fill = PatternFill("solid", fgColor="0B0B0C")
    for c in ws[head_row]:
        c.font = Font(bold=True, color="E0B100")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    for r in data["rows"]:
        ws.append([r["employee_id"], r["name"], r["department"], r["position"],
                   r["present"], r["tolerance"], r["late"], r["absent"], r["attended"],
                   r["workdays"], r["rate"]])
    t = data["totals"]
    ws.append([])
    total_row = ["", "TOTAL", "", "", t["present"], t["tolerance"], t["late"], t["absent"],
                 t["attended"], "", ""]
    ws.append(total_row)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    widths = [14, 26, 18, 20, 11, 10, 11, 11, 11, 10, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Sheet 2: detail harian dengan jam masuk & pulang
    detail = data.get("detail") or []
    ws2 = wb.create_sheet("Detail Harian")
    ws2.append(["Detail Harian — Jam Masuk & Pulang"])
    ws2.append([f"Periode: {data['start']} s/d {data['end']}"])
    ws2.append([])
    dheaders = ["Tanggal", "ID Karyawan", "Nama", "Departemen", "Jam Masuk", "Jam Pulang", "Status"]
    ws2.append(dheaders)
    dhead_row = ws2.max_row
    for c in ws2[dhead_row]:
        c.font = Font(bold=True, color="E0B100")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    for d in detail:
        ws2.append([d["date"], d["employee_id"], d["name"], d["department"],
                    d["check_in"], d["check_out"], d["status_label"]])
    for i, w in enumerate([14, 14, 26, 18, 12, 12, 16], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_bytes(data: dict) -> bytes:
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "Rekap Absensi Karyawan - MitraKeuangan", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Periode: {data['start']} s/d {data['end']}    Hari kerja: {data['workdays']}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    cols = [("No", 10), ("ID", 20), ("Nama", 48), ("Departemen", 34), ("Tepat", 16),
            ("Tol", 14), ("Telat", 16), ("Absen", 16), ("Total", 16), ("Hr Kerja", 18), ("%", 14)]
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(11, 11, 12)
    pdf.set_text_color(224, 177, 0)
    for h, w in cols:
        pdf.cell(w, 7, h, border=1, fill=True, align="C")
    pdf.ln(7)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for i, r in enumerate(data["rows"], 1):
        vals = [str(i), _latin(r["employee_id"]), _latin(r["name"])[:30],
                _latin(r.get("department") or "-")[:22], str(r["present"]), str(r["tolerance"]),
                str(r["late"]), str(r["absent"]), str(r["attended"]), str(r["workdays"]),
                str(r["rate"])]
        for (h, w), v in zip(cols, vals):
            pdf.cell(w, 6, v, border=1, align="C" if h not in ("Nama", "Departemen", "ID") else "L")
        pdf.ln(6)
    t = data["totals"]
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(10 + 20 + 48 + 34, 7, "TOTAL", border=1, align="R")
    for v, w in [(t["present"], 16), (t["tolerance"], 14), (t["late"], 16),
                 (t["absent"], 16), (t["attended"], 16)]:
        pdf.cell(w, 7, str(v), border=1, align="C")
    pdf.cell(18, 7, "", border=1)
    pdf.cell(14, 7, "", border=1)

    # Halaman detail harian (jam masuk & pulang)
    detail = data.get("detail") or []
    if detail:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Detail Harian - Jam Masuk & Pulang", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 6, f"Periode: {data['start']} s/d {data['end']}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        dcols = [("Tanggal", 28), ("ID", 24), ("Nama", 58), ("Departemen", 40),
                 ("Masuk", 22), ("Pulang", 22), ("Status", 34)]
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(11, 11, 12)
        pdf.set_text_color(224, 177, 0)
        for h, w in dcols:
            pdf.cell(w, 7, h, border=1, fill=True, align="C")
        pdf.ln(7)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)
        for d in detail:
            vals = [str(d["date"]), _latin(d["employee_id"]), _latin(d["name"])[:36],
                    _latin(d.get("department") or "-")[:24], d["check_in"], d["check_out"],
                    _latin(d["status_label"])]
            for (h, w), v in zip(dcols, vals):
                pdf.cell(w, 6, v, border=1, align="L" if h in ("Nama", "Departemen", "ID") else "C")
            pdf.ln(6)

    out = pdf.output()
    return bytes(out)


@api_router.get("/attendance/summary/export")
async def export_summary(format: str = Query("xlsx"), start: str = Query(...), end: str = Query(...),
                         user_id: Optional[str] = Query(None), q: Optional[str] = Query(None),
                         user: dict = Depends(require_roles(MONITOR_ROLES))):
    data = await _build_summary(start, end, user_id, q, include_detail=True)
    fname = f"rekap-absensi-{start}_sd_{end}"
    if format == "pdf":
        content = _pdf_bytes(data)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})
    content = _xlsx_bytes(data)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})
def _summarize_today(users: list, rec_by_user: dict) -> dict:
    present = tolerance = late = 0
    dept_map = {}
    for u in users:
        r = rec_by_user.get(str(u["_id"]))
        st = r["status"] if r else "absent"
        if st == "present":
            present += 1
        elif st == "tolerance":
            tolerance += 1
        elif st == "late":
            late += 1
        dept = u.get("department") or "Umum"
        dept_map.setdefault(dept, {"department": dept, "total": 0, "hadir": 0})
        dept_map[dept]["total"] += 1
        if st in ("present", "tolerance", "late"):
            dept_map[dept]["hadir"] += 1
    return {"present": present, "tolerance": tolerance, "late": late,
            "departments": list(dept_map.values())}


async def _seven_day_trend() -> list:
    days = [datetime.now(TZ) - timedelta(days=i) for i in range(6, -1, -1)]
    dates = [d.strftime("%Y-%m-%d") for d in days]
    recs = await db.attendance.find({"date": {"$in": dates}}).to_list(5000)
    counts = {dt: {"present": 0, "late": 0} for dt in dates}
    for r in recs:
        c = counts.get(r.get("date"))
        if not c:
            continue
        st = r.get("status")
        if st in ("present", "tolerance"):
            c["present"] += 1
        elif st == "late":
            c["late"] += 1
    trend = []
    for day, dt in zip(days, dates):
        p, l = counts[dt]["present"], counts[dt]["late"]
        trend.append({"date": dt, "label": day.strftime("%a"), "present": p, "late": l, "total": p + l})
    return trend


def _recent_activity(recs: list, users: list) -> list:
    recent = sorted([r for r in recs if r.get("check_in")],
                    key=lambda r: r["check_in"], reverse=True)[:8]
    user_by_id = {str(u["_id"]): u for u in users}
    return [att_public(r, user_by_id.get(r["user_id"])) for r in recent]


@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(require_roles(MONITOR_ROLES))):
    date = today_str()
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    total = len(users)
    recs = await db.attendance.find({"date": date}).to_list(2000)
    rec_by_user = {r["user_id"]: r for r in recs}

    summary = _summarize_today(users, rec_by_user)
    present, tolerance, late = summary["present"], summary["tolerance"], summary["late"]
    absent = total - present - tolerance - late

    return {
        "date": date,
        "total": total, "present": present, "tolerance": tolerance, "late": late, "absent": absent,
        "attendance_rate": round((present + tolerance + late) / total * 100, 1) if total else 0,
        "departments": summary["departments"],
        "trend": await _seven_day_trend(),
        "recent": _recent_activity(recs, users),
    }


@api_router.get("/settings")
async def read_settings(user: dict = Depends(get_current_user)):
    return await get_settings()


@api_router.put("/settings")
async def update_settings(body: SettingsBody, user: dict = Depends(require_roles(HR_ROLES))):
    _validate_hhmm(body.work_start)
    _validate_hhmm(body.work_end)
    if not (0 <= body.tolerance_minutes <= 120):
        raise HTTPException(status_code=400, detail="Toleransi harus 0–120 menit")
    payload = {"work_start": body.work_start, "work_end": body.work_end,
               "tolerance_minutes": body.tolerance_minutes,
               "leave_quota": body.leave_quota if body.leave_quota is not None else 12}
    if user["role"] in APPROVER_ROLES:
        await _apply_settings(payload)
        return {"ok": True, "applied": True, **payload}
    summary = (f"Ubah jadwal kerja: masuk {body.work_start}, pulang {body.work_end}, "
               f"toleransi {body.tolerance_minutes} menit")
    return await _create_request("settings", user, summary=summary, payload=payload)


# ----------------------------------------------------------------------------
# Pengajuan Cuti (Leave) — approval 3 layer: HRD -> Direksi/Manager -> Reviewer
# ----------------------------------------------------------------------------
LEAVE_TYPE_LABEL = {"tahunan": "Cuti Tahunan", "sakit": "Sakit", "izin": "Izin"}
LEAVE_STAGES = ["hrd", "direksi", "reviewer"]
LEAVE_STAGE_LABEL = {"hrd": "HRD", "direksi": "Direksi/Manager", "reviewer": "Reviewer", "done": "Selesai"}


def _count_weekdays_inclusive(start: str, end: str) -> int:
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end, "%Y-%m-%d").date()
    if e < s:
        return 0
    n, d = 0, s
    while d <= e:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


async def _leave_used_days(user_id: str, year: int) -> int:
    total = 0
    async for r in db.leave_requests.find(
            {"user_id": user_id, "status": "approved", "leave_type": "tahunan",
             "start_date": {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}}):
        total += int(r.get("days", 0))
    return total


def leave_public(r: dict) -> dict:
    return {
        "id": str(r["_id"]),
        "user_id": r.get("user_id"),
        "user_name": r.get("user_name"),
        "department": r.get("department"),
        "leave_type": r.get("leave_type"),
        "leave_type_label": LEAVE_TYPE_LABEL.get(r.get("leave_type"), r.get("leave_type")),
        "start_date": r.get("start_date"),
        "end_date": r.get("end_date"),
        "days": r.get("days"),
        "reason": r.get("reason"),
        "status": r.get("status"),
        "stage": r.get("stage"),
        "stage_label": LEAVE_STAGE_LABEL.get(r.get("stage"), r.get("stage")),
        "approvals": r.get("approvals", []),
        "reject_reason": r.get("reject_reason"),
        "created_at": r.get("created_at"),
    }


def _leave_can_act(user: dict, req: dict) -> bool:
    if req.get("status") != "pending":
        return False
    role = user.get("role")
    if role == "owner":
        return True
    stage = req.get("stage")
    if stage == "hrd":
        return role == "hrd"
    if stage == "direksi":
        return role in ("direksi", "manager")
    if stage == "reviewer":
        return bool(user.get("is_reviewer"))
    return False


async def _leave_stage_notify_ids(stage: str) -> list:
    if stage == "hrd":
        q = {"role": "hrd"}
    elif stage == "direksi":
        q = {"role": {"$in": ["direksi", "manager"]}}
    elif stage == "reviewer":
        q = {"is_reviewer": True}
    else:
        return []
    docs = await db.users.find(q, {"_id": 1}).to_list(200)
    ids = [str(d["_id"]) for d in docs]
    owners = await db.users.find({"role": "owner"}, {"_id": 1}).to_list(50)
    ids += [str(d["_id"]) for d in owners]
    return list(set(ids))


@api_router.get("/leave/balance")
async def leave_balance(user_id: Optional[str] = Query(None),
                        user: dict = Depends(get_current_user)):
    target = user_id if (user_id and user["role"] in MONITOR_ROLES) else str(user["_id"])
    settings = await get_settings()
    quota = settings.get("leave_quota", 12)
    year = datetime.now(TZ).year
    used = await _leave_used_days(target, year)
    pend = 0
    async for r in db.leave_requests.find(
            {"user_id": target, "status": "pending", "leave_type": "tahunan",
             "start_date": {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}}):
        pend += int(r.get("days", 0))
    return {"year": year, "quota": quota, "used": used, "pending": pend,
            "remaining": max(0, quota - used)}


@api_router.post("/leave-requests")
async def create_leave_request(body: LeaveRequestBody, user: dict = Depends(get_current_user)):
    if body.leave_type not in LEAVE_TYPE_LABEL:
        raise HTTPException(status_code=400, detail="Jenis cuti tidak valid")
    try:
        s = datetime.strptime(body.start_date, "%Y-%m-%d").date()
        e = datetime.strptime(body.end_date, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Tanggal tidak valid")
    if e < s:
        raise HTTPException(status_code=400, detail="Tanggal selesai sebelum tanggal mulai")
    days = _count_weekdays_inclusive(body.start_date, body.end_date)
    if days <= 0:
        raise HTTPException(status_code=400, detail="Rentang tidak mengandung hari kerja (Senin–Jumat)")
    settings = await get_settings()
    quota = settings.get("leave_quota", 12)
    if body.leave_type == "tahunan":
        used = await _leave_used_days(str(user["_id"]), s.year)
        if used + days > quota:
            raise HTTPException(status_code=400,
                                detail=f"Sisa cuti tahunan tidak cukup (sisa {max(0, quota - used)} hari)")
    doc = {
        "user_id": str(user["_id"]), "user_name": user.get("name"),
        "department": user.get("department"), "leave_type": body.leave_type,
        "start_date": body.start_date, "end_date": body.end_date, "days": days,
        "reason": body.reason or "", "status": "pending", "stage": "hrd",
        "approvals": [], "created_at": now_iso(),
    }
    res = await db.leave_requests.insert_one(doc)
    doc["_id"] = res.inserted_id
    await send_push_to_users(await _leave_stage_notify_ids("hrd"), "Pengajuan Cuti Baru",
                             f"{user.get('name')} mengajukan {LEAVE_TYPE_LABEL[body.leave_type]} "
                             f"{days} hari ({body.start_date} s/d {body.end_date}).",
                             url="/#leave", tag="leave")
    return leave_public(doc)


@api_router.get("/leave-requests")
async def list_leave_requests(scope: str = Query("all"), status: Optional[str] = Query(None),
                              user: dict = Depends(get_current_user)):
    can_approve = user["role"] in MONITOR_ROLES or bool(user.get("is_reviewer"))
    query = {}
    if scope == "mine" or not can_approve:
        query["user_id"] = str(user["_id"])
    if status:
        query["status"] = status
    docs = await db.leave_requests.find(query).sort("created_at", -1).to_list(2000)
    items = []
    for d in docs:
        it = leave_public(d)
        it["can_act"] = _leave_can_act(user, d)
        items.append(it)
    return {"items": items, "count": len(items)}


@api_router.get("/leave-requests/pending-count")
async def leave_pending_count(user: dict = Depends(get_current_user)):
    docs = await db.leave_requests.find({"status": "pending"}).to_list(2000)
    n = sum(1 for d in docs if _leave_can_act(user, d))
    return {"count": n}


@api_router.post("/leave-requests/{req_id}/approve")
async def approve_leave(req_id: str, body: LeaveDecisionBody, user: dict = Depends(get_current_user)):
    req = await db.leave_requests.find_one({"_id": _oid(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Pengajuan tidak ditemukan")
    if not _leave_can_act(user, req):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang menyetujui tahap ini")
    stage = req["stage"]
    approvals = req.get("approvals", [])
    approvals.append({"stage": stage, "by_id": str(user["_id"]), "by_name": user.get("name"),
                      "decision": "approved", "note": body.note or "", "at": now_iso()})
    idx = LEAVE_STAGES.index(stage)
    if idx + 1 < len(LEAVE_STAGES):
        next_stage = LEAVE_STAGES[idx + 1]
        await db.leave_requests.update_one(
            {"_id": req["_id"]}, {"$set": {"stage": next_stage, "approvals": approvals}})
        await send_push_to_users(await _leave_stage_notify_ids(next_stage), "Persetujuan Cuti",
                                 f"Pengajuan cuti {req.get('user_name')} menunggu persetujuan Anda "
                                 f"(tahap {LEAVE_STAGE_LABEL[next_stage]}).",
                                 url="/#leave", tag="leave")
        return {"ok": True, "stage": next_stage, "status": "pending"}
    await db.leave_requests.update_one(
        {"_id": req["_id"]}, {"$set": {"stage": "done", "status": "approved", "approvals": approvals}})
    await send_push_to_users([req["user_id"]], "Cuti Disetujui",
                             f"Pengajuan {LEAVE_TYPE_LABEL.get(req.get('leave_type'))} "
                             f"{req.get('start_date')} s/d {req.get('end_date')} telah DISETUJUI penuh.",
                             url="/#leave", tag="leave")
    return {"ok": True, "stage": "done", "status": "approved"}


@api_router.post("/leave-requests/{req_id}/reject")
async def reject_leave(req_id: str, body: LeaveDecisionBody, user: dict = Depends(get_current_user)):
    req = await db.leave_requests.find_one({"_id": _oid(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Pengajuan tidak ditemukan")
    if not _leave_can_act(user, req):
        raise HTTPException(status_code=403, detail="Anda tidak berwenang menolak tahap ini")
    approvals = req.get("approvals", [])
    approvals.append({"stage": req["stage"], "by_id": str(user["_id"]), "by_name": user.get("name"),
                      "decision": "rejected", "note": body.note or "", "at": now_iso()})
    await db.leave_requests.update_one(
        {"_id": req["_id"]}, {"$set": {"status": "rejected", "stage": "done", "approvals": approvals,
                                       "reject_reason": body.note or "Ditolak"}})
    await send_push_to_users([req["user_id"]], "Cuti Ditolak",
                             f"Pengajuan cuti Anda ditolak pada tahap {LEAVE_STAGE_LABEL.get(req['stage'], req['stage'])}. "
                             f"{body.note or ''}".strip(), url="/#leave", tag="leave")
    return {"ok": True, "status": "rejected"}


@api_router.put("/employees/{emp_id}/reviewer")
async def set_reviewer(emp_id: str, body: ReviewerBody, user: dict = Depends(require_roles(HR_ROLES))):
    target = await db.users.find_one({"_id": _oid(emp_id)})
    if not target:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    if body.is_reviewer and target.get("role") != "manager":
        raise HTTPException(status_code=400, detail="Reviewer hanya dapat ditunjuk untuk Manager")
    await db.users.update_one({"_id": target["_id"]}, {"$set": {"is_reviewer": body.is_reviewer}})
    return {"ok": True, "is_reviewer": body.is_reviewer}


# ----------------------------------------------------------------------------
# Employees (management) — with HRD approval workflow
# ----------------------------------------------------------------------------
def req_public(r: dict) -> dict:
    payload = r.get("payload")
    if isinstance(payload, dict):
        payload = {k: v for k, v in payload.items() if k != "password_hash"}
    return {
        "id": str(r["_id"]),
        "action": r["action"],
        "summary": r.get("summary"),
        "status": r["status"],
        "target_emp_id": r.get("target_emp_id"),
        "payload": payload,
        "requested_by_name": r.get("requested_by_name"),
        "created_at": r.get("created_at"),
        "reviewed_by_name": r.get("reviewed_by_name"),
        "reviewed_at": r.get("reviewed_at"),
        "reject_reason": r.get("reject_reason"),
    }


async def _apply_create_employee(payload: dict) -> dict:
    if await db.users.find_one({"email": payload["email"]}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    count = await db.users.count_documents({})
    doc = {
        "name": payload["name"], "email": payload["email"],
        "password_hash": payload["password_hash"],
        "role": payload["role"], "department": payload.get("department"),
        "position": payload.get("position"), "phone": payload.get("phone"),
        "employee_id": f"EMP-{count + 1:04d}", "face_descriptor": None,
        "created_at": now_iso(),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    return public_user(doc)


async def _apply_update_employee(emp_id: str, update: dict) -> dict:
    if update:
        await db.users.update_one({"_id": ObjectId(emp_id)}, {"$set": update})
    u = await db.users.find_one({"_id": ObjectId(emp_id)})
    return public_user(u)


async def _apply_delete_employee(emp_id: str) -> None:
    await db.users.delete_one({"_id": ObjectId(emp_id)})
    await db.attendance.delete_many({"user_id": emp_id})


async def _create_request(action: str, requester: dict, summary: str,
                          payload: Optional[dict] = None, target_emp_id: Optional[str] = None) -> dict:
    doc = {
        "action": action, "payload": payload, "target_emp_id": target_emp_id,
        "summary": summary, "status": "pending",
        "requested_by": str(requester["_id"]), "requested_by_name": requester.get("name"),
        "created_at": now_iso(),
        "reviewed_by": None, "reviewed_by_name": None, "reviewed_at": None, "reject_reason": None,
    }
    res = await db.employee_requests.insert_one(doc)
    doc["_id"] = res.inserted_id
    await send_push_to_users(await _approver_ids(), "Permintaan Baru", summary,
                             url="/#approvals", tag="request")
    out = req_public(doc)
    out["pending"] = True
    return out


def _guard_hrd_role(requester: dict, role: Optional[str]):
    """HRD tidak boleh menetapkan/ mengubah seseorang menjadi owner/direksi."""
    if requester["role"] == "hrd" and role is not None and role not in HRD_ASSIGNABLE_ROLES:
        raise HTTPException(status_code=403, detail="HRD tidak dapat menetapkan role Owner/Direksi")


@api_router.get("/employees")
async def list_employees(user: dict = Depends(require_roles(MONITOR_ROLES))):
    users = await db.users.find({}, {"password_hash": 0}).sort("created_at", 1).to_list(1000)
    return [public_user(u) for u in users]


@api_router.post("/employees")
async def create_employee(body: EmployeeBody, user: dict = Depends(require_roles(HR_ROLES))):
    email = body.email.lower()
    role = body.role if body.role in VALID_ROLES else "staff"
    _guard_hrd_role(user, role)
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    payload = {
        "name": body.name, "email": email, "password_hash": hash_password(body.password),
        "role": role, "department": body.department, "position": body.position, "phone": body.phone,
    }
    if user["role"] in APPROVER_ROLES:
        return await _apply_create_employee(payload)
    return await _create_request("create", user, payload=payload,
                                 summary=f"Tambah karyawan: {body.name} ({email}) — role {role}")


@api_router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, body: EmployeeUpdate, user: dict = Depends(require_roles(HR_ROLES))):
    _guard_hrd_role(user, body.role)
    target = await db.users.find_one({"_id": ObjectId(emp_id)})
    if not target:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    if user["role"] == "hrd" and target.get("role") in {"owner", "direksi"}:
        raise HTTPException(status_code=403, detail="HRD tidak dapat mengubah data Owner/Direksi")
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if "password" in update:
        update["password_hash"] = hash_password(update.pop("password"))
    if user["role"] in APPROVER_ROLES:
        return await _apply_update_employee(emp_id, update)
    return await _create_request("update", user, payload=update, target_emp_id=emp_id,
                                 summary=f"Ubah data karyawan: {target.get('name')}")


@api_router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, user: dict = Depends(require_roles(HR_ROLES))):
    if str(user["_id"]) == emp_id:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun sendiri")
    target = await db.users.find_one({"_id": ObjectId(emp_id)})
    if not target:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    if user["role"] == "hrd" and target.get("role") in {"owner", "direksi"}:
        raise HTTPException(status_code=403, detail="HRD tidak dapat menghapus Owner/Direksi")
    if user["role"] in APPROVER_ROLES:
        await _apply_delete_employee(emp_id)
        return {"ok": True}
    return await _create_request("delete", user, target_emp_id=emp_id,
                                 summary=f"Hapus karyawan: {target.get('name')}")


# ----------------------------------------------------------------------------
# Employee change requests (HRD -> Direksi approval)
# ----------------------------------------------------------------------------
@api_router.get("/employee-requests")
async def list_employee_requests(status: Optional[str] = Query(None),
                                 user: dict = Depends(require_roles(HR_ROLES))):
    q = {}
    if status:
        q["status"] = status
    if user["role"] not in APPROVER_ROLES:
        q["requested_by"] = str(user["_id"])  # HRD hanya melihat permintaannya sendiri
    reqs = await db.employee_requests.find(q).sort("created_at", -1).to_list(500)
    return [req_public(r) for r in reqs]


@api_router.get("/employee-requests/pending-count")
async def pending_count(user: dict = Depends(require_roles(APPROVER_ROLES))):
    n = await db.employee_requests.count_documents({"status": "pending"})
    return {"count": n}


@api_router.post("/employee-requests/{req_id}/approve")
async def approve_request(req_id: str, user: dict = Depends(require_roles(APPROVER_ROLES))):
    req = await db.employee_requests.find_one({"_id": ObjectId(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Permintaan tidak ditemukan")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Permintaan sudah diproses")
    action = req["action"]
    result = None
    if action == "create":
        result = await _apply_create_employee(req["payload"])
    elif action == "update":
        result = await _apply_update_employee(req["target_emp_id"], req.get("payload") or {})
    elif action == "delete":
        await _apply_delete_employee(req["target_emp_id"])
    elif action == "settings":
        await _apply_settings(req.get("payload") or {})
    await db.employee_requests.update_one({"_id": req["_id"]}, {"$set": {
        "status": "approved", "reviewed_by": str(user["_id"]),
        "reviewed_by_name": user.get("name"), "reviewed_at": now_iso()}})
    await send_push_to_users([req["requested_by"]], "Permintaan Disetujui",
                             req.get("summary"), url="/#approvals", tag="request")
    return {"ok": True, "result": result}


@api_router.post("/employee-requests/{req_id}/reject")
async def reject_request(req_id: str, body: RejectBody, user: dict = Depends(require_roles(APPROVER_ROLES))):
    req = await db.employee_requests.find_one({"_id": ObjectId(req_id)})
    if not req:
        raise HTTPException(status_code=404, detail="Permintaan tidak ditemukan")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Permintaan sudah diproses")
    await db.employee_requests.update_one({"_id": req["_id"]}, {"$set": {
        "status": "rejected", "reviewed_by": str(user["_id"]),
        "reviewed_by_name": user.get("name"), "reviewed_at": now_iso(),
        "reject_reason": body.reason or "Ditolak"}})
    await send_push_to_users([req["requested_by"]], "Permintaan Ditolak",
                             f"{req.get('summary')} — {body.reason or 'Ditolak'}",
                             url="/#approvals", tag="request")
    return {"ok": True}


# ----------------------------------------------------------------------------
# Notifications (role-aware)
# ----------------------------------------------------------------------------
_ACTION_LABEL = {"create": "Tambah karyawan", "update": "Ubah data karyawan",
                 "delete": "Hapus karyawan", "settings": "Ubah jadwal kerja"}


@api_router.get("/notifications")
async def notifications(user: dict = Depends(get_current_user)):
    role = user.get("role")
    items = []
    if role in APPROVER_ROLES:
        reqs = await db.employee_requests.find({"status": "pending"}).sort("created_at", -1).to_list(50)
        for r in reqs:
            items.append({
                "id": str(r["_id"]), "type": "pending", "route": "approvals",
                "title": f"Permintaan: {_ACTION_LABEL.get(r['action'], r['action'])}",
                "body": r.get("summary"), "by": r.get("requested_by_name"),
                "created_at": r.get("created_at"),
            })
    elif role == "hrd":
        reqs = await db.employee_requests.find(
            {"requested_by": str(user["_id"]), "status": {"$in": ["approved", "rejected"]}}
        ).sort("reviewed_at", -1).to_list(50)
        for r in reqs:
            items.append({
                "id": str(r["_id"]), "type": r["status"], "route": "approvals",
                "title": "Permintaan disetujui" if r["status"] == "approved" else "Permintaan ditolak",
                "body": r.get("summary"), "by": r.get("reviewed_by_name"),
                "created_at": r.get("reviewed_at"),
            })

    # Leave (cuti) notifications — for anyone who can act on a stage
    pending_leaves = await db.leave_requests.find({"status": "pending"}).sort("created_at", -1).to_list(100)
    for r in pending_leaves:
        if _leave_can_act(user, r):
            items.append({
                "id": "leave-" + str(r["_id"]), "type": "leave", "route": "leave",
                "title": f"Persetujuan Cuti — {LEAVE_STAGE_LABEL.get(r.get('stage'), '')}",
                "body": f"{r.get('user_name')} · {LEAVE_TYPE_LABEL.get(r.get('leave_type'), '')} {r.get('days')} hari",
                "by": r.get("user_name"), "created_at": r.get("created_at"),
            })
    my_leaves = await db.leave_requests.find(
        {"user_id": str(user["_id"]), "status": {"$in": ["approved", "rejected"]}}
    ).sort("created_at", -1).to_list(20)
    for r in my_leaves:
        items.append({
            "id": "leave-" + str(r["_id"]), "type": r["status"], "route": "leave",
            "title": "Cuti disetujui" if r["status"] == "approved" else "Cuti ditolak",
            "body": f"{LEAVE_TYPE_LABEL.get(r.get('leave_type'), '')} {r.get('start_date')} s/d {r.get('end_date')}",
            "by": "", "created_at": r.get("created_at"),
        })

    items.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return {"items": items, "count": len(items)}


# ----------------------------------------------------------------------------
# Web Push subscriptions
# ----------------------------------------------------------------------------
@api_router.get("/push/vapid-public-key")
async def vapid_public_key():
    return {"key": VAPID_PUBLIC_KEY, "configured": bool(VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY)}


@api_router.post("/push/subscribe")
async def push_subscribe(body: PushSubscribeBody, user: dict = Depends(get_current_user)):
    endpoint = body.subscription.get("endpoint")
    if not endpoint:
        raise HTTPException(status_code=400, detail="Subscription tidak valid")
    await db.push_subscriptions.update_one(
        {"endpoint": endpoint},
        {"$set": {"user_id": str(user["_id"]), "subscription": body.subscription,
                  "endpoint": endpoint, "updated_at": now_iso()},
         "$setOnInsert": {"created_at": now_iso()}},
        upsert=True)
    return {"ok": True}


@api_router.post("/push/unsubscribe")
async def push_unsubscribe(body: PushUnsubscribeBody, user: dict = Depends(get_current_user)):
    if body.endpoint:
        await db.push_subscriptions.delete_one(
            {"endpoint": body.endpoint, "user_id": str(user["_id"])})
    return {"ok": True}


@api_router.get("/push/status")
async def push_status(user: dict = Depends(get_current_user)):
    n = await db.push_subscriptions.count_documents({"user_id": str(user["_id"])})
    return {"subscribed": n > 0, "count": n,
            "configured": bool(VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY)}


# ----------------------------------------------------------------------------
# App wiring + seed
# ----------------------------------------------------------------------------
app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


DEMO_USERS = [
    {"name": "Budi Santoso", "email": None, "role": "owner",
     "department": "Direksi", "position": "Owner / CEO"},
    {"name": "Siti Rahmawati", "email": "direksi@company.com", "role": "direksi",
     "department": "Direksi", "position": "Direktur Operasional"},
    {"name": "Agus Wijaya", "email": "manager@company.com", "role": "manager",
     "department": "Penjualan", "position": "Manager Penjualan"},
    {"name": "Maya Anggraini", "email": "hrd@company.com", "role": "hrd",
     "department": "HRD", "position": "Manager HRD"},
    {"name": "Dewi Lestari", "email": "dewi@company.com", "role": "staff",
     "department": "Penjualan", "position": "Sales Executive"},
    {"name": "Rizki Pratama", "email": "rizki@company.com", "role": "staff",
     "department": "Keuangan", "position": "Staff Keuangan"},
    {"name": "Nina Putri", "email": "nina@company.com", "role": "staff",
     "department": "HRD", "position": "Staff HRD"},
    {"name": "Eko Saputra", "email": "eko@company.com", "role": "staff",
     "department": "Operasional", "position": "Teknisi"},
]

TEST_CREDENTIALS_DOC = """# Test Credentials

All demo accounts use password: **password123**

| Role | Email | Password |
|------|-------|----------|
| Owner | owner@company.com | password123 |
| Direksi | direksi@company.com | password123 |
| Manager | manager@company.com | password123 |
| HRD | hrd@company.com | password123 |
| Staff | dewi@company.com | password123 |
| Staff | rizki@company.com | password123 |
| Staff | nina@company.com | password123 |
| Staff | eko@company.com | password123 |

## Auth endpoints
- POST /api/auth/login  body {email, password} -> {token, user}
- POST /api/auth/register
- GET  /api/auth/me  (Bearer token)

## Notes
- Frontend stores token in localStorage and sends `Authorization: Bearer <token>`.
- Monitoring (dashboard/employees/all-attendance): roles owner, direksi, manager, hrd.
- Direct management (apply employees & offices): roles owner, direksi.
- HRD can request employee create/update/delete (cannot touch owner/direksi roles); requests need owner/direksi approval via /api/employee-requests/{id}/approve.
- Default office: "Kantor Pusat Jakarta" lat -6.208763 lng 106.845599 radius 250m, QR: OFFICE-HQJKT00001
"""


async def _ensure_indexes():
    await db.users.create_index("email", unique=True)
    await db.attendance.create_index([("user_id", 1), ("date", 1)])
    await db.push_subscriptions.create_index("endpoint", unique=True)
    await db.push_subscriptions.create_index("user_id")


async def _seed_users(admin_email: str, admin_password: str, demo: bool = True):
    # In production (demo=False) only the Owner admin account is created.
    users = DEMO_USERS if demo else DEMO_USERS[:1]
    for i, d in enumerate(users):
        email = admin_email if d["email"] is None else d["email"]
        pwd = admin_password if email == admin_email else "password123"
        existing = await db.users.find_one({"email": email})
        if existing is None:
            await db.users.insert_one({
                **{k: v for k, v in d.items() if k != "email"}, "email": email,
                "password_hash": hash_password(pwd),
                "phone": None, "employee_id": f"EMP-{i + 1:04d}",
                "face_descriptor": None, "created_at": now_iso(),
            })
        elif not verify_password(pwd, existing["password_hash"]):
            await db.users.update_one({"email": email},
                                      {"$set": {"password_hash": hash_password(pwd)}})


async def _seed_office():
    if await db.offices.count_documents({}) == 0:
        await db.offices.insert_one({
            "name": "Kantor Pusat Jakarta", "lat": -6.208763, "lng": 106.845599,
            "radius_m": 250, "qr_code": "OFFICE-HQJKT00001",
        })


def _rand_below(n: int) -> int:
    return secrets.randbelow(n)


async def _seed_attendance_history():
    """Populate the last 7 days of demo attendance for charts (uses `secrets`)."""
    if await db.attendance.count_documents({}) != 0:
        return
    users = await db.users.find({"role": {"$in": ["manager", "staff", "direksi"]}}).to_list(100)
    for i in range(7, 0, -1):
        day = datetime.now(TZ) - timedelta(days=i)
        date = day.strftime("%Y-%m-%d")
        for u in users:
            if _rand_below(100) < 12:
                continue  # absent
            late = _rand_below(100) < 20
            hour = 9 if late else 8
            minute = (5 + _rand_below(36)) if late else _rand_below(56)
            ci = day.replace(hour=hour, minute=minute, second=0, microsecond=0).astimezone(timezone.utc)
            co = day.replace(hour=17, minute=_rand_below(60), second=0, microsecond=0).astimezone(timezone.utc)
            await db.attendance.insert_one({
                "user_id": str(u["_id"]), "date": date,
                "check_in": ci.isoformat(), "check_out": co.isoformat(),
                "status": "late" if late else "present",
                "method": secrets.choice(["face", "qr"]),
                "office_name": "Kantor Pusat Jakarta", "distance_m": 5 + _rand_below(116),
                "within_geofence": True, "lat": -6.2087, "lng": 106.8456,
            })


def _write_test_credentials():
    try:
        Path("/app/memory/test_credentials.md").write_text(TEST_CREDENTIALS_DOC)
    except Exception as e:
        logger.warning(f"could not write creds: {e}")


async def seed():
    await _ensure_indexes()
    await _seed_users(os.environ["ADMIN_EMAIL"], os.environ["ADMIN_PASSWORD"], demo=SEED_DEMO)
    if SEED_DEMO:
        await _seed_office()
        await _seed_attendance_history()
        _write_test_credentials()
        logger.info("Seed complete (DEMO mode: demo accounts & sample data created).")
    else:
        logger.info("Seed complete (PRODUCTION mode: only Owner admin account created).")


@app.on_event("startup")
async def on_startup():
    await seed()
    try:
        hh, mm = map(int, REMINDER_TIME.split(":"))
        scheduler.add_job(_send_attendance_reminders,
                          CronTrigger(hour=hh, minute=mm),
                          id="attendance_reminder", replace_existing=True)
        if not scheduler.running:
            scheduler.start()
        logger.info(f"Scheduler started; daily reminder at {REMINDER_TIME} {TZ}.")
    except Exception as e:
        logger.warning(f"scheduler init failed: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        if scheduler.running:
            scheduler.shutdown(wait=False)
    except Exception:
        pass
    client.close()
